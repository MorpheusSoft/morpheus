import openpyxl
import os

path = "/home/lzambrano/Downloads/LP A&B N° 20 Vigente 02.07.2026.xlsx"
if not os.path.exists(path):
    print(f"File {path} does not exist!")
else:
    print(f"File exists. Size: {os.path.getsize(path)} bytes")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        print("Sheets in workbook:")
        print(wb.sheetnames)
        sheet = wb.active
        print(f"Active sheet: {sheet.title}")
        print(f"Max row: {sheet.max_row}, Max column: {sheet.max_column}")
        
        # Print first 10 rows
        print("\nFirst 10 rows:")
        for r in range(1, 11):
            row_vals = [cell.value for cell in sheet[r]]
            print(f"Row {r}: {row_vals}")
    except Exception as e:
        print(f"Error reading file: {e}")
